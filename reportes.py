import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import subprocess
import io
import urllib.request
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

# --- CONFIGURACIÓN DE VERSIÓN ---
VERSION_ACTUAL = "v1.0.9" 
EXE_NAME = "ReportesSecurithor.exe"
REPO_API_URL = "https://api.github.com/repos/Miniacidus/Report-Securithor/releases/latest"
URL_DOWNLOAD = "https://github.com/Miniacidus/Report-Securithor/releases/latest/download/ReportesSecurithor.exe"

ruta_mensual = ""
ruta_anual = ""
FILAS_POR_PAGINA = 50 

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}

# --- SISTEMA DE ACTUALIZACIÓN (CON LIMPIEZA _MEI PARA EVITAR ERROR DLL) ---
def actualizar_programa():
    try:
        btn_update.config(text="Buscando...", state="disabled")
        root.update()
        req = urllib.request.Request(REPO_API_URL, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req) as response:
            data = json.loads(response.read().decode())
            ultima_version = data['tag_name']

        if ultima_version == VERSION_ACTUAL:
            messagebox.showinfo("Software al día", f"Ya tienes la versión más reciente ({VERSION_ACTUAL}).")
            return

        if messagebox.askyesno("Nueva Versión", f"Actualización disponible: {ultima_version}\n¿Descargar e instalar?"):
            temp_exe = "Update_Nuevo.exe"
            urllib.request.urlretrieve(URL_DOWNLOAD, temp_exe)
            
            # Script .bat robusto: Espera 20s y borra carpetas temporales corruptas
            with open("update_script.bat", "w") as f:
                f.write('@echo off\n')
                f.write('title Actualizando y Limpiando Sistema...\n')
                f.write('echo 1. Esperando cierre total del programa (20s)...\n')
                f.write('timeout /t 20 /nobreak > nul\n')
                
                f.write('echo 2. Eliminando residuos temporales (Fix DLL error)...\n')
                # Este comando es vital: borra la basura de Python anterior
                f.write('for /d %%i in ("%temp%\\_MEI*") do rd /s /q "%%i" 2>nul\n')
                
                f.write('echo 3. Instalando nueva version...\n')
                f.write(f'del "{EXE_NAME}"\n')
                f.write(f'ren "{temp_exe}" "{EXE_NAME}"\n')
                f.write(f'start "" "{EXE_NAME}"\n')
                f.write('del "%~f0"\n')
            
            messagebox.showinfo("Actualizando", "El programa se cerrará para actualizarse y limpiarse.")
            subprocess.Popen(["update_script.bat"], shell=True)
            root.destroy()
    except Exception as e:
        messagebox.showerror("Error de red", f"No se pudo actualizar: {e}")
    finally:
        try: btn_update.config(text="🔄 Check for Updates", state="normal")
        except: pass

# --- GESTIÓN DE ARCHIVOS EXTERNOS (BAJAS Y NOMBRES) ---
def asegurar_archivos_config():
    if not os.path.exists("bajas.txt"):
        with open("bajas.txt", "w", encoding="utf-8") as f: f.write("")
    if not os.path.exists("nombres.txt"):
        with open("nombres.txt", "w", encoding="utf-8") as f: f.write("Cuenta, Nombre")

def abrir_txt(archivo):
    asegurar_archivos_config()
    try: subprocess.Popen(["notepad.exe", archivo])
    except: messagebox.showerror("Error", f"No se pudo abrir {archivo}")

def cargar_bajas():
    if not os.path.exists("bajas.txt"): return []
    with open("bajas.txt", "r", encoding="utf-8") as f: return [line.strip() for line in f if line.strip()]

def cargar_diccionario_nombres():
    nombres = {}
    if os.path.exists("nombres.txt"):
        with open("nombres.txt", "r", encoding="utf-8") as f:
            for line in f:
                if "," in line:
                    partes = line.strip().split(",", 1)
                    nombres[partes[0].strip()] = partes[1].strip()
    return nombres

# --- MOTOR DE PROCESAMIENTO CSV (ROBUSTO) ---
def corregir_fecha(texto):
    if pd.isna(texto): return pd.NaT
    texto = str(texto).lower().replace('p. m.', 'pm').replace('a. m.', 'am').replace('p.m.', 'pm').replace('a.m.', 'am')
    try: return pd.to_datetime(texto, dayfirst=True)
    except: return pd.NaT

def leer_csv_robusto(ruta):
    try:
        # Intenta UTF-8 primero, luego Latin-1 si falla
        with open(ruta, 'r', encoding='utf-8-sig', errors='replace') as f: contenido = f.read()
        if len(contenido) < 10:
             with open(ruta, 'r', encoding='latin-1', errors='replace') as f: contenido = f.read()
        
        df = pd.read_csv(io.StringIO(contenido), header=3)
        df.columns = df.columns.str.strip()
        
        # Busca columna de fecha flexiblemente
        posibles = ['Llegada', 'Recibido', 'Fecha', 'Arrival', 'Received']
        col_f = next((c for c in posibles if c in df.columns), None)
        if not col_f: raise Exception("No se encontró columna de fecha válida.")
        
        df = df.rename(columns={col_f: 'Llegada'})
        df['FechaCompleta'] = df['Llegada'].apply(corregir_fecha)
        df = df.dropna(subset=['FechaCompleta'])
        
        # Limpia número de cuenta (quita .0)
        df['Cuenta'] = df['Cuenta'].astype(str).str.replace(r'\.0$', '', regex=True)
        df['Cuenta_Num'] = pd.to_numeric(df['Cuenta'], errors='coerce').fillna(0)
        
        # Filtra Bajas y Mapa Nombres
        bajas = cargar_bajas(); df = df[~df['Cuenta'].isin(bajas)]
        mapa = cargar_diccionario_nombres(); df['Nombre del Cliente'] = df['Cuenta'].map(mapa).fillna('')
        return df
    except Exception as e: raise Exception(f"Error procesando CSV: {e}")

def aplicar_formato_excel(ruta_excel, modo="asistencia"):
    wb = load_workbook(ruta_excel)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        color_fondo = "4472C4" 
        if sheet_name == "Sin_Reportar": color_fondo = "FFC000" 
        if sheet_name == "Solo_Sucesos": color_fondo = "9BC2E6" 

        # Configuración de Página para Impresión
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.sheet_properties.pageSetUpPr.fitToPage = True 
        ws.page_setup.fitToWidth = 1 # Forza ancho a 1 página
        ws.page_setup.fitToHeight = 0 
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
        ws.print_title_rows = '1:1' # Repite encabezados

        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row == 1:
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                    cell.fill = PatternFill("solid", fgColor=color_fondo)
        
        # Ajuste de anchos de columna
        ws.column_dimensions['A'].width = 12 
        ws.column_dimensions['B'].width = 42
        if modo == "asistencia":
            for col in range(3, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 5.5
        
        # Saltos de página
        for i in range(FILAS_POR_PAGINA + 2, ws.max_row + 1, FILAS_POR_PAGINA):
            ws.row_breaks.append(Break(id=i-1))
    wb.save(ruta_excel)

def generar_reportes():
    global ruta_mensual, ruta_anual
    if not ruta_mensual and not ruta_anual:
        messagebox.showwarning("Atención", "Selecciona al menos un archivo CSV.")
        return
    try: 
        btn_generar.config(text="⏳ Procesando...", state="disabled")
        root.update()
        
        # --- REPORTE 1: MENSUAL DE ASISTENCIA ---
        if ruta_mensual:
            df = leer_csv_robusto(ruta_mensual)
            fecha_max = df['FechaCompleta'].max()
            df['Dia'] = df['FechaCompleta'].dt.day
            df['Hora'] = df['FechaCompleta'].dt.strftime('%H:%M')
            pivot = df.pivot_table(index=['Cuenta_Num', 'Cuenta', 'Nombre del Cliente'], columns='Dia', values='Hora', aggfunc='max').fillna('/')
            out = os.path.join(os.path.dirname(ruta_mensual), f"1_Reporte_Asistencia_{MESES_ES[fecha_max.month]}.xlsx")
            pivot.to_excel(out)
            aplicar_formato_excel(out, "asistencia")
            
        # --- REPORTE 2: ANUAL DE FALLAS ---
        if ruta_anual:
            df_a = leer_csv_robusto(ruta_anual)
            fecha_ref_a = df_a['FechaCompleta'].max()
            u = df_a.groupby(['Cuenta_Num', 'Cuenta', 'Nombre del Cliente'])['FechaCompleta'].max().reset_index()
            u['Dias'] = (fecha_ref_a - u['FechaCompleta']).dt.days
            f = u[u['Dias'] > 4].copy() # Filtro de 4 días
            
            # Detecta código 88 para diferenciar Fallas de Solo Sucesos
            def detectar_test(cta):
                # Busca '88' en cualquier columna de esa cuenta (Alarma, Evento, etc)
                has_88 = df_a[df_a['Cuenta'] == cta].astype(str).to_string().find('88') != -1
                return "" if has_88 else "Solo sucesos"
                
            f['Notas'] = f['Cuenta'].apply(detectar_test)
            f['Desde'] = (f['FechaCompleta'] + pd.Timedelta(days=1)).dt.strftime('%d/%m/%Y')
            f_ord = f.sort_values('Cuenta_Num')

            # Selecciona SOLO columnas limpias para el reporte final
            df_sin = f_ord[f_ord['Notas'] == ''][['Cuenta', 'Nombre del Cliente', 'Desde', 'Notas']]
            df_suc = f_ord[f_ord['Notas'] == 'Solo sucesos'][['Cuenta', 'Nombre del Cliente', 'Desde', 'Notas']]
            
            out_a = os.path.join(os.path.dirname(ruta_anual), f"2_Reporte_Fallas_Anual.xlsx")
            with pd.ExcelWriter(out_a, engine='openpyxl') as writer:
                df_sin.to_excel(writer, sheet_name='Sin_Reportar', index=False)
                df_suc.to_excel(writer, sheet_name='Solo_Sucesos', index=False)
            aplicar_formato_excel(out_a, "fallas")

        messagebox.showinfo("¡Éxito!", "Reportes generados correctamente.")
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un problema: {e}")
    finally:
        btn_generar.config(text="GENERATE REPORTS", state="normal")

# --- INTERFAZ GRÁFICA ---
def sel_m(): 
    global ruta_mensual
    r = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    if r: ruta_mensual = r; lbl_m.config(text=os.path.basename(r), fg="green")

def sel_a(): 
    global ruta_anual
    r = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    if r: ruta_anual = r; lbl_a.config(text=os.path.basename(r), fg="green")

root = tk.Tk()
root.title(f"Securithor Automator {VERSION_ACTUAL}")
root.geometry("450x500")

tk.Label(root, text="System Report Generator", font=("Arial", 14, "bold")).pack(pady=10)
f1 = tk.LabelFrame(root, text="Mensual", padx=10, pady=5); f1.pack(pady=5, fill="x")
tk.Button(f1, text="Select CSV", command=sel_m).pack(); lbl_m = tk.Label(f1, text="(vacio)", fg="gray"); lbl_m.pack()

f2 = tk.LabelFrame(root, text="Anual", padx=10, pady=5); f2.pack(pady=5, fill="x")
tk.Button(f2, text="Select CSV", command=sel_a).pack(); lbl_a = tk.Label(f2, text="(vacio)", fg="gray"); lbl_a.pack()

fr = tk.Frame(root); fr.pack(pady=10)
tk.Button(fr, text="📝 Bajas", command=lambda: abrir_txt("bajas.txt")).grid(row=0, column=0, padx=5)
tk.Button(fr, text="👤 Nombres", command=lambda: abrir_txt("nombres.txt")).grid(row=0, column=1, padx=5)

btn_update = tk.Button(root, text="🔄 Check for Updates", command=actualizar_programa)
btn_update.pack(pady=5)
btn_generar = tk.Button(root, text="GENERATE REPORTS", command=generar_reportes, bg="green", fg="white", font=("Arial", 11, "bold"), height=2)
btn_generar.pack(pady=15)

asegurar_archivos_config()
root.mainloop()



